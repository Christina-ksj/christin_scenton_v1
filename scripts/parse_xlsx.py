#!/usr/bin/env python3
"""
parse_xlsx.py — GSC xlsx 파일을 파싱해서 gsc_data.json 생성
Usage: python3 parse_xlsx.py --perf <실적.xlsx> --index <색인.xlsx> --out data/gsc_data.json
"""
import argparse
import json
import re
import sys
from pathlib import Path


def require_openpyxl():
    try:
        import openpyxl
        return openpyxl
    except ImportError:
        print("✗ openpyxl 미설치. 다음 명령으로 설치하세요: pip3 install openpyxl")
        sys.exit(1)


def get_sheet(wb, candidates):
    """여러 후보 시트명 중 첫 번째 매칭되는 시트 반환"""
    names_lower = {n.lower(): n for n in wb.sheetnames}
    for c in candidates:
        if c.lower() in names_lower:
            return wb[names_lower[c.lower()]]
    # 순서 기반 fallback
    if wb.sheetnames:
        return wb[wb.sheetnames[0]]
    return None


def sheet_to_rows(sheet):
    """시트를 헤더+데이터 rows로 변환"""
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return [], []
    # 헤더 찾기 (첫 번째 비어있지 않은 행)
    header_idx = 0
    for i, row in enumerate(rows):
        if any(cell is not None for cell in row):
            header_idx = i
            break
    headers = [str(h).strip() if h is not None else f"col{i}" for i, h in enumerate(rows[header_idx])]
    data = []
    for row in rows[header_idx + 1:]:
        if not any(cell is not None for cell in row):
            continue
        record = {headers[i]: row[i] for i in range(min(len(headers), len(row)))}
        data.append(record)
    return headers, data


def normalize_ctr(val):
    """CTR 값 정규화 (% 문자열 → float 0~1)"""
    if val is None:
        return 0.0
    if isinstance(val, (int, float)):
        return float(val) if val <= 1.0 else float(val) / 100.0
    s = str(val).replace('%', '').strip()
    try:
        f = float(s)
        return f / 100.0 if f > 1.0 else f
    except ValueError:
        return 0.0


def normalize_pos(val):
    if val is None:
        return 0.0
    try:
        return round(float(val), 2)
    except (ValueError, TypeError):
        return 0.0


def parse_perf_xlsx(path: str) -> dict:
    """실적 xlsx 파싱"""
    openpyxl = require_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  실적 시트 목록: {wb.sheetnames}")

    result = {
        "chart": [],
        "keywords": [],
        "pages": [],
        "devices": [],
        "countries": [],
        "opportunity": [],
        "low_ctr": [],
    }

    # ── 날짜별 성과 (차트) ──
    date_sheet = get_sheet(wb, ["날짜", "Date", "Dates", "Overview", "요약", "일별"])
    if date_sheet:
        _, rows = sheet_to_rows(date_sheet)
        for r in rows:
            date_val = None
            clicks = 0
            impressions = 0
            ctr = 0.0
            position = 0.0
            for k, v in r.items():
                kl = k.lower()
                if any(x in kl for x in ["날짜", "date"]):
                    if v is not None:
                        date_val = str(v)[:10]
                elif any(x in kl for x in ["클릭", "click"]):
                    try:
                        clicks = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif any(x in kl for x in ["노출", "impress"]):
                    try:
                        impressions = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif "ctr" in kl:
                    ctr = normalize_ctr(v)
                elif any(x in kl for x in ["순위", "position"]):
                    position = normalize_pos(v)
            if date_val and re.match(r"\d{4}-\d{2}-\d{2}", date_val):
                result["chart"].append({
                    "날짜": date_val,
                    "클릭수": clicks,
                    "노출": impressions,
                    "CTR": round(ctr, 4),
                    "순위": position,
                })
        result["chart"].sort(key=lambda x: x["날짜"])

    # ── 키워드 ──
    kw_sheet = get_sheet(wb, ["쿼리", "Queries", "Keywords", "검색어", "인기 검색어"])
    if kw_sheet:
        _, rows = sheet_to_rows(kw_sheet)
        for r in rows:
            query = None
            clicks = impressions = 0
            ctr = pos = 0.0
            for k, v in r.items():
                kl = k.lower()
                if any(x in kl for x in ["쿼리", "query", "검색어", "키워드", "keyword"]):
                    query = str(v).strip() if v else None
                elif any(x in kl for x in ["클릭", "click"]):
                    try:
                        clicks = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif any(x in kl for x in ["노출", "impress"]):
                    try:
                        impressions = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif "ctr" in kl:
                    ctr = normalize_ctr(v)
                elif any(x in kl for x in ["순위", "position"]):
                    pos = normalize_pos(v)
            if query:
                result["keywords"].append({
                    "인기 검색어": query,
                    "클릭수": clicks,
                    "노출": impressions,
                    "CTR": round(ctr, 4),
                    "게재 순위": pos,
                })

    # ── 페이지 ──
    page_sheet = get_sheet(wb, ["페이지", "Pages", "URLs"])
    if page_sheet:
        _, rows = sheet_to_rows(page_sheet)
        for r in rows:
            page = None
            clicks = impressions = 0
            ctr = pos = 0.0
            for k, v in r.items():
                kl = k.lower()
                if any(x in kl for x in ["페이지", "page", "url"]):
                    page = str(v).strip() if v else None
                elif any(x in kl for x in ["클릭", "click"]):
                    try:
                        clicks = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif any(x in kl for x in ["노출", "impress"]):
                    try:
                        impressions = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif "ctr" in kl:
                    ctr = normalize_ctr(v)
                elif any(x in kl for x in ["순위", "position"]):
                    pos = normalize_pos(v)
            if page:
                result["pages"].append({
                    "인기 페이지": page,
                    "클릭수": clicks,
                    "노출": impressions,
                    "CTR": round(ctr, 4),
                    "게재 순위": pos,
                })

    # ── 기기 ──
    dev_sheet = get_sheet(wb, ["기기", "Devices", "Device"])
    if dev_sheet:
        _, rows = sheet_to_rows(dev_sheet)
        for r in rows:
            device = None
            clicks = impressions = 0
            for k, v in r.items():
                kl = k.lower()
                if any(x in kl for x in ["기기", "device"]):
                    device = str(v).strip() if v else None
                elif any(x in kl for x in ["클릭", "click"]):
                    try:
                        clicks = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif any(x in kl for x in ["노출", "impress"]):
                    try:
                        impressions = int(v or 0)
                    except (ValueError, TypeError):
                        pass
            if device:
                result["devices"].append({"기기": device, "클릭수": clicks, "노출": impressions})

    # ── 국가 ──
    country_sheet = get_sheet(wb, ["국가", "Countries", "Country"])
    if country_sheet:
        _, rows = sheet_to_rows(country_sheet)
        for r in rows:
            country = None
            clicks = impressions = 0
            for k, v in r.items():
                kl = k.lower()
                if any(x in kl for x in ["국가", "country"]):
                    country = str(v).strip() if v else None
                elif any(x in kl for x in ["클릭", "click"]):
                    try:
                        clicks = int(v or 0)
                    except (ValueError, TypeError):
                        pass
                elif any(x in kl for x in ["노출", "impress"]):
                    try:
                        impressions = int(v or 0)
                    except (ValueError, TypeError):
                        pass
            if country:
                result["countries"].append({"국가": country, "클릭수": clicks, "노출": impressions})

    # ── 기회 키워드 (순위 11~20, 노출 높음) ──
    all_kw = result["keywords"]
    result["opportunity"] = sorted(
        [k for k in all_kw if 10 < k["게재 순위"] <= 20 and k["노출"] > 50],
        key=lambda x: -x["노출"]
    )[:10]

    # ── 저CTR 키워드 ──
    result["low_ctr"] = sorted(
        [k for k in all_kw if k["노출"] >= 200 and k["CTR"] < 0.01],
        key=lambda x: -x["노출"]
    )[:10]

    return result


def parse_index_xlsx(path: str) -> dict:
    """색인 xlsx 파싱"""
    openpyxl = require_openpyxl()
    wb = openpyxl.load_workbook(path, data_only=True)
    print(f"  색인 시트 목록: {wb.sheetnames}")

    index_result = {
        "indexed": 0,
        "not_indexed": 0,
        "issues": [],
    }

    # 모든 시트를 탐색해서 색인/미색인 수치 찾기
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        _, rows = sheet_to_rows(sheet)
        for r in rows:
            reason = None
            source = None
            count = 0
            for k, v in r.items():
                kl = k.lower()
                if any(x in kl for x in ["사유", "reason", "이유", "상태", "status", "유형", "type"]):
                    reason = str(v).strip() if v else None
                elif any(x in kl for x in ["소스", "source"]):
                    source = str(v).strip() if v else None
                elif any(x in kl for x in ["페이지", "page", "url", "수", "count"]):
                    try:
                        count = int(v or 0)
                    except (ValueError, TypeError):
                        pass

            if reason and count > 0:
                r_lower = reason.lower()
                if any(x in r_lower for x in ["색인됨", "indexed", "검색됨"]) and "미" not in r_lower and "not" not in r_lower:
                    index_result["indexed"] = count
                elif any(x in r_lower for x in ["미색인", "not indexed", "색인 안됨", "색인이 생성되지 않음"]):
                    index_result["not_indexed"] += count
                    index_result["issues"].append({
                        "사유": reason,
                        "소스": source or "웹사이트",
                        "페이지": count,
                    })
                elif reason not in ("", "사유"):
                    index_result["issues"].append({
                        "사유": reason,
                        "소스": source or "웹사이트",
                        "페이지": count,
                    })

    # 미색인 합계 재계산
    if index_result["not_indexed"] == 0 and index_result["issues"]:
        index_result["not_indexed"] = sum(i["페이지"] for i in index_result["issues"])

    return index_result


def main():
    parser = argparse.ArgumentParser(description="GSC xlsx → JSON 변환기")
    parser.add_argument("--perf", required=True, help="실적 xlsx 파일 경로")
    parser.add_argument("--index", default="", help="색인 xlsx 파일 경로 (옵션)")
    parser.add_argument("--out", default="data/gsc_data.json", help="출력 JSON 경로")
    args = parser.parse_args()

    print("\n[parse_xlsx] GSC 데이터 파싱 시작")

    if not Path(args.perf).exists():
        print(f"✗ 파일 없음: {args.perf}")
        sys.exit(1)

    # 실적 파싱
    data = parse_perf_xlsx(args.perf)
    print(f"  차트: {len(data['chart'])}일, 키워드: {len(data['keywords'])}개, 페이지: {len(data['pages'])}개")

    # 색인 파싱
    if args.index and Path(args.index).exists():
        index_data = parse_index_xlsx(args.index)
        data["index"] = index_data
        print(f"  색인: {index_data['indexed']:,} / 미색인: {index_data['not_indexed']:,}")
    else:
        print("  색인 파일 없음 — 기존 데이터 유지")
        data["index"] = None

    # 저장
    out_path = Path(args.out)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"✓ 저장 완료: {out_path}")


if __name__ == "__main__":
    main()
